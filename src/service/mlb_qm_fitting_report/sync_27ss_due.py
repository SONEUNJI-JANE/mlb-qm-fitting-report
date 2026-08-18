import json

SEASON = "27SS"
DUE_OVERRIDE_SETTING_KEY = "mlb_qm_27ss_due_overrides"
SHEET_NAME_HINT = "27SS DATA"
HEADER_ROW = 3  # 1-based: DS/QUARTER/... 헤더가 있는 행
# 0-based column index (엑셀 O/P/R/S열)
COL_STYLE = 4
COL_QC_DUE = 14   # O: QC CFM DUE (FIT)
COL_PP_DUE = 15   # P: PP CFM DUE 1차
COL_TOP_DUE = 17  # R: TOP CFM DUE 1차
COL_TOP_DUE_2 = 18  # S: TOP CFM DUE 2차


def _iso(value) -> str | None:
    if value is None or value == "":
        return None
    return str(value)[:10]


def read_due_overrides(xlsx_path: str) -> dict:
    """엑셀 '▶27SS DATA' 탭 O/P/R/S열 -> {style_code: {"qc_due","pp_due","top_due","top_due_2"}}.

    소싱팀 관리 파일. 27SS 납기 대비 로직이 여기서 바뀌었고, styles 테이블의 qc_due/pp_due/top_due는
    다른 시스템(PLM, qc_due_src 등 _src 컬럼으로 추적)이 주기적으로 다시 덮어쓸 수 있어서 직접
    upsert하지 않는다 — 대신 이 값을 별도로 저장해두고 live_app이 styles 값보다 우선해서 쓴다."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        sheet_name = next(s for s in wb.sheetnames if SHEET_NAME_HINT in s)
        ws = wb[sheet_name]

        overrides = {}
        for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            style_code = row[COL_STYLE]
            if not style_code:
                continue
            overrides[style_code] = {
                "qc_due": _iso(row[COL_QC_DUE]),
                "pp_due": _iso(row[COL_PP_DUE]),
                "top_due": _iso(row[COL_TOP_DUE]),
                "top_due_2": _iso(row[COL_TOP_DUE_2]),
            }
        return overrides
    finally:
        wb.close()


def sync_27ss_due(settings: dict) -> dict:
    from src.service.mlb_qm_fitting_report.supabase_client import upsert_setting

    xlsx_path = settings["due_date_sources"][SEASON]
    overrides = read_due_overrides(xlsx_path)
    upsert_setting(settings, DUE_OVERRIDE_SETTING_KEY, json.dumps(overrides))
    return {"styles": len(overrides)}


if __name__ == "__main__":
    from src.service.mlb_qm_fitting_report.config import load_settings

    settings = load_settings()
    result = sync_27ss_due(settings)
    print(f"synced: {result}")
