import shutil
import tempfile
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

_SHEET = "FIT TRACK CHART"
_DATA_START_ROW = 9

_COL_STYLE = 3
_COL_ITEM = 1
_COL_QUARTER = 2
_COL_VENDOR = 8
_COL_TD = 11
_COL_QA = 13
_COL_ETD_KR = 15
_COL_ETD_CN = 16
_COL_ETD_GL = 17
_COL_FIT_DONE = 101
_COL_PP_DONE = 102
_COL_TOP_DONE = 103

# 회차별 블록: [접수, Fitting, 전달, status, 사유] 5칸 1묶음. 1차부터 순서대로 보고 접수일이
# 채워진 마지막 회차를 "현재 진행 회차"로 본다. 보정(X~AG)은 대시보드 3단계(FIT/PP/TOP) 집계엔 안
# 잡히지만, FIT이 아직 시작 전일 때 "이전 단계" 상세(상태/전달일/사유)로는 보여준다.
_PIPELINE_STAGES = ["보정", "FIT", "PP", "TOP"]
_ROUND_BLOCKS = {
    "보정": ["X", "AC"],
    "FIT": ["AJ", "AO", "AT", "AY", "BD"],
    "PP": ["BL", "BQ", "BV", "CA"],
    "TOP": ["CF", "CK", "CP"],
}
_ROUND_LABELS = ["1ST", "2ND", "3RD", "4TH", "5TH"]

_DUE_SHEET = "DUE_DATA(2)"
_DUE_DATA_START_ROW = 4
_DUE_COL_STYLE = 5
_DUE_COL_LABEL = 10
_DUE_COL_QC_DUE = 18
_DUE_COL_PP_DUE = 20
_DUE_COL_TOP_DUE = 21

# "가장 빠른 ETD로부터 (-)일" 기준표. 구분(KNIT/SWEATER/WOVEN/DENIM)+워시기준+수량기준 라벨별
# QC(=FIT)/PP/TOP 오프셋 일수. DUE_DATA(2) 시트에 해당 스타일 행 자체가 없거나 ETD/생산LT가
# 비어 QC DUE/PP DUE/TOP DUE가 깨진(1899년대) 값으로 나올 때만 이 표로 직접 계산한다.
LABEL_OFFSETS = {
    "KNIT워싱일반": {"category": "KNIT", "wash": "워싱", "qty_tier": "일반", "FIT": 75, "PP": 50, "TOP": 10},
    "KNIT워싱3천장이하": {"category": "KNIT", "wash": "워싱", "qty_tier": "3천장이하", "FIT": 75, "PP": 40, "TOP": 10},
    "KNIT논워싱일반": {"category": "KNIT", "wash": "논워싱", "qty_tier": "일반", "FIT": 75, "PP": 50, "TOP": 10},
    "KNIT논워싱3천장이하": {"category": "KNIT", "wash": "논워싱", "qty_tier": "3천장이하", "FIT": 75, "PP": 40, "TOP": 10},
    "SWEATER워싱일반": {"category": "SWEATER", "wash": "워싱", "qty_tier": "일반", "FIT": 75, "PP": 50, "TOP": 10},
    "SWEATER워싱3천장이하": {"category": "SWEATER", "wash": "워싱", "qty_tier": "3천장이하", "FIT": 75, "PP": 40, "TOP": 10},
    "SWEATER논워싱일반": {"category": "SWEATER", "wash": "논워싱", "qty_tier": "일반", "FIT": 75, "PP": 50, "TOP": 10},
    "SWEATER논워싱3천장이하": {"category": "SWEATER", "wash": "논워싱", "qty_tier": "3천장이하", "FIT": 75, "PP": 40, "TOP": 10},
    "WOVEN워싱일반": {"category": "WOVEN", "wash": "워싱", "qty_tier": "일반", "FIT": 75, "PP": 60, "TOP": 21},
    "WOVEN논워싱일반": {"category": "WOVEN", "wash": "논워싱", "qty_tier": "일반", "FIT": 75, "PP": 55, "TOP": 17},
    "DENIM워싱일반": {"category": "DENIM", "wash": "워싱", "qty_tier": "일반", "FIT": 100, "PP": 70, "TOP": 45},
}

# RAW(Apparel)의 item(2글자 코드) -> 구분(KNIT/SWEATER/WOVEN/DENIM). DUE_DATA(2)에 없는 스타일도
# 이걸로 구분을 알아내서 LABEL_OFFSETS를 적용할 수 있게 한다.
ITEM_CATEGORY = {
    "DJ": "WOVEN", "PD": "WOVEN", "DV": "WOVEN", "FD": "WOVEN", "WJ": "WOVEN", "JP": "WOVEN",
    "JK": "WOVEN", "VT": "WOVEN", "WS": "WOVEN", "WP": "WOVEN", "SM": "WOVEN", "SK": "WOVEN", "FU": "WOVEN",
    "DR": "DENIM", "DD": "DENIM", "DP": "DENIM", "DS": "DENIM", "DK": "DENIM",
    "SW": "KNIT", "LG": "KNIT", "SP": "KNIT", "TP": "KNIT", "PT": "KNIT", "BS": "KNIT", "HD": "KNIT",
    "MT": "KNIT", "RS": "KNIT", "TR": "KNIT", "OP": "KNIT", "TS": "KNIT", "TK": "KNIT", "TO": "KNIT",
    "PQ": "KNIT", "KS": "KNIT",
    "KP": "SWEATER", "KC": "SWEATER", "KB": "SWEATER",
}
_QTY_TIER_THRESHOLD = 3000  # AN열(Qty TTL)이 이 값 이하면 "3천장이하"


class HeaderMismatchError(Exception):
    pass


def _check_headers(ws, header_row: int, expected: dict[int, str]) -> None:
    """엑셀 서식이 바뀌어 하드코딩된 컬럼 인덱스가 어긋나면, 잘못된 값을 조용히 집계하는 대신
    바로 죽여서 알린다. expected: {1-based 컬럼번호: 헤더에 포함돼야 할 부분 문자열}."""
    for col, substr in expected.items():
        actual = ws.cell(row=header_row, column=col).value
        if substr not in str(actual or ""):
            letter = get_column_letter(col)
            raise HeaderMismatchError(
                f"[{ws.title}] {letter}{header_row}열 헤더가 예상과 다릅니다. "
                f"기대: '{substr}' 포함 / 실제: '{actual}'. 엑셀 서식이 바뀌었는지 확인하세요."
            )


def _find_header_row(ws, col: int, substr: str, max_row: int = 10) -> int:
    """같은 시트라도 파일 사본마다 위쪽에 삽입된 행 수가 달라 헤더 행 번호가 밀릴 수 있어서
    (예: RAW(Apparel) 시트가 사본에 따라 4행 또는 5행), 고정 행 번호 대신 이 컬럼에서 헤더
    텍스트를 찾아 실제 행을 반환한다. max_row 안에 못 찾으면 바로 죽여서 알린다."""
    for row in range(1, max_row + 1):
        if substr in str(ws.cell(row=row, column=col).value or ""):
            return row
    letter = get_column_letter(col)
    raise HeaderMismatchError(
        f"[{ws.title}] {letter}열에서 '{substr}' 헤더를 {max_row}행 안에서 못 찾았습니다. "
        f"엑셀 서식이 바뀌었는지 확인하세요."
    )


def _as_date(value):
    return value.date() if isinstance(value, datetime) else None


def _valid_due(value):
    d = _as_date(value)
    if d is None or d.year < 2000:
        return None
    return d


def _read_due_data_sheet(wb) -> dict:
    """STYLE -> {"label": str|None, "due": {"FIT": date|None, "PP": date|None, "TOP": date|None}}"""
    ws = wb[_DUE_SHEET]
    _check_headers(ws, _DUE_DATA_START_ROW - 1, {
        _DUE_COL_STYLE: "STYLE",
        _DUE_COL_LABEL: "라벨",
        _DUE_COL_QC_DUE: "QC DUE",
        _DUE_COL_PP_DUE: "PP DUE",
        _DUE_COL_TOP_DUE: "TOP DUE",
    })
    result = {}
    for r in ws.iter_rows(min_row=_DUE_DATA_START_ROW, values_only=True):
        style = r[_DUE_COL_STYLE - 1]
        if not style or not isinstance(style, str) or style in result:
            continue
        result[style] = {
            "label": r[_DUE_COL_LABEL - 1],
            "due": {
                "FIT": _valid_due(r[_DUE_COL_QC_DUE - 1]),
                "PP": _valid_due(r[_DUE_COL_PP_DUE - 1]),
                "TOP": _valid_due(r[_DUE_COL_TOP_DUE - 1]),
            },
        }
    return result


def _earliest_etd(row) -> date | None:
    dates = [d for d in (_as_date(row[_COL_ETD_KR]), _as_date(row[_COL_ETD_CN]), _as_date(row[_COL_ETD_GL])) if d]
    return min(dates) if dates else None


def _round_block_cols(start_letter: str) -> dict:
    start = column_index_from_string(start_letter)
    return {"received": start, "fitting": start + 1, "delivered": start + 2, "status": start + 3, "reason": start + 4}


_ROUND_BLOCK_COLS = {stage: [_round_block_cols(b) for b in blocks] for stage, blocks in _ROUND_BLOCKS.items()}


def _first_received(row, stage: str):
    """그 stage 1회차부터 순서대로 봐서 접수일(날짜 타입)이 실제로 찍힌 첫 회차의 접수일.
    승인 이후 "다음 스테이지 접수까지 며칠 걸렸나"(스테이지 전환 리드타임)를 재려면 그 스테이지가
    실제로 시작된 시점이 필요해서, 최신 회차가 아니라 맨 처음 회차 기준으로 따로 잡는다."""
    for cols in _ROUND_BLOCK_COLS[stage]:
        received = row[cols["received"] - 1]
        if isinstance(received, datetime):
            return received.date()
    return None


def _all_rounds(row, stage: str) -> list[dict]:
    """그 stage에서 실제 데이터(접수일/전달일/status) 있는 회차들을 순서대로 전부 반환.
    회차 간 소요일수(1ST FIT→2ND FIT 등)를 재려면 최신/첫 회차만으론 안 되고 전체가 필요하다."""
    blocks = _ROUND_BLOCK_COLS[stage]
    rounds = []
    for i, cols in enumerate(blocks):
        received = row[cols["received"] - 1]
        delivered = row[cols["delivered"] - 1]
        status = row[cols["status"] - 1]
        if not (isinstance(received, datetime) or isinstance(delivered, datetime) or status):
            continue
        confirm_date = _as_date(delivered)
        if confirm_date is None and status == "Approved":
            confirm_date = _as_date(row[cols["fitting"] - 1])
        rounds.append({
            "round": _ROUND_LABELS[i],
            "received": _as_date(received),
            "status": status,
            "confirm_date": confirm_date,
        })
    return rounds


def _latest_round(row, stage: str) -> dict:
    """그 stage에서 실제 데이터(접수일/전달일/status) 있는 마지막 회차의 status/전달일/사유.
    접수 칸엔 "전 회차 승인 완료" 같은 안내 텍스트만, fitting 칸엔 "넥목업" 같은 메모만 들어있고
    실제 날짜/status가 없는 회차는 회차로 치지 않는다(이전 회차에서 이미 끝났다는 표시일 뿐이라
    이 텍스트들 때문에 실제 데이터 있는 이전 회차를 덮어쓰면 안 됨). fitting 칸은 날짜 아닌 메모가
    들어갈 수 있어서 회차 판단 기준에서 뺐다 — 접수일/전달일(둘 다 날짜 타입)이나 status만 본다."""
    blocks = _ROUND_BLOCK_COLS[stage]
    last = None
    for i, cols in enumerate(blocks):
        received = row[cols["received"] - 1]
        delivered = row[cols["delivered"] - 1]
        if isinstance(received, datetime) or isinstance(delivered, datetime) or row[cols["status"] - 1]:
            last = i
    if last is None:
        return {"round": None, "status": None, "confirm_date": None, "reason": None, "first_received": None}
    cols = blocks[last]
    status = row[cols["status"] - 1]
    # 전달일(확정일) 칸이 비어있어도 status가 Approved면 fitting일(실제 피팅 진행일)을 대신 쓴다
    # — 담당자가 상태만 적고 확정일 기입을 빠뜨린 경우가 있어서, 완전히 날짜 없이 두는 것보다 낫다.
    confirm_date = _as_date(row[cols["delivered"] - 1])
    if confirm_date is None and status == "Approved":
        confirm_date = _as_date(row[cols["fitting"] - 1])
    return {
        "round": _ROUND_LABELS[last],
        "status": status,
        "confirm_date": confirm_date,
        "reason": row[cols["reason"] - 1],
        "first_received": _first_received(row, stage),
    }


def _stage_details(row) -> dict:
    """보정/FIT/PP/TOP 각 단계의 현재 진행 회차 상세. 보정은 대시보드 집계 대상은 아니지만, FIT이
    아직 시작 전일 때 "이전 단계" 상세(상태/전달일/사유)로 보여주려고 같이 계산해둔다.
    보정에서 FIT 안 거치고 바로 PP로 넘어간 경우 FIT은 '접수 전'이 아니라 '생략'으로 구분한다
    (PP에 접수 기록이 있으면 그렇게 판단)."""
    details = {stage: _latest_round(row, stage) for stage in _PIPELINE_STAGES}
    for stage in _PIPELINE_STAGES:
        details[stage]["rounds"] = _all_rounds(row, stage)
    if details["FIT"]["round"] is None and details["PP"]["round"] is not None:
        details["FIT"]["status"] = "생략(보정→PP 직행)"
    return details


def _read_fit_track_sheet(wb) -> dict:
    """STYLE -> {"etd": date|None, "fit_done"/"pp_done"/"top_done": bool, "detail": {...}}. FIT TRACK
    CHART에 없는 스타일(아직 fitting 안 시작)은 이 dict에 없고, 호출하는 쪽에서 done=False/etd=None으로 처리한다."""
    ws = wb[_SHEET]
    _check_headers(ws, _DATA_START_ROW - 1, {
        _COL_STYLE + 1: "Style NO",
        _COL_VENDOR + 1: "Vendor",
        _COL_TD + 1: "TD",
        _COL_QA + 1: "QA",
        _COL_ETD_KR + 1: "한국납기",
        _COL_ETD_CN + 1: "중국납기",
        _COL_ETD_GL + 1: "글로벌",
        _COL_FIT_DONE + 1: "FIT완료",
        _COL_PP_DONE + 1: "PP완료",
        _COL_TOP_DONE + 1: "TOP 완료",
    })
    result = {}
    for r in ws.iter_rows(min_row=_DATA_START_ROW, values_only=True):
        style_code = r[_COL_STYLE]
        if not style_code or not isinstance(style_code, str) or style_code in result:
            continue
        result[style_code] = {
            "etd": _earliest_etd(r),
            "vendor": r[_COL_VENDOR],
            "item": r[_COL_ITEM],
            "quarter": r[_COL_QUARTER],
            "td": r[_COL_TD],
            "qa": r[_COL_QA],
            "fit_done": r[_COL_FIT_DONE] == 1,
            "pp_done": r[_COL_PP_DONE] == 1,
            "top_done": r[_COL_TOP_DONE] == 1,
            "detail": _stage_details(r),
        }
    return result


_RAW_SHEET = "RAW(Apparel)"
_RAW_COL_STYLE = column_index_from_string("Q")
_RAW_COL_ITEM = column_index_from_string("D")
_RAW_COL_MODEL_COUNT = column_index_from_string("AI")  # 모델수. 1인 행만 그 스타일의 대표 1건으로 센다.
_RAW_COL_STATUS = column_index_from_string("AS")  # GO/DROP. DROP은 제외.
_RAW_COL_QTY_TTL = column_index_from_string("AN")
_RAW_COL_WASHED = column_index_from_string("BU")
_RAW_COL_RDD_KR = column_index_from_string("AU")
_RAW_COL_RDD_CN = column_index_from_string("AX")
_RAW_COL_ETA_GL = column_index_from_string("BA")


def compute_label(item_code, washed_raw, qty_ttl) -> str | None:
    """LABEL_OFFSETS 표(사용자 제공 기준표)엔 수량 구간(3천장이하)이 KNIT/SWEATER에만 있고
    WOVEN/DENIM은 물량 상관없이 "일반" 한 줄뿐이라, 그 두 구분에서만 수량 구간을 따진다."""
    category = ITEM_CATEGORY.get(item_code)
    if not category:
        return None
    wash = "논워싱" if (not washed_raw or str(washed_raw).strip().upper() == "X") else "워싱"
    if category in ("KNIT", "SWEATER") and isinstance(qty_ttl, (int, float)) and qty_ttl <= _QTY_TIER_THRESHOLD:
        qty_tier = "3천장이하"
    else:
        qty_tier = "일반"
    return f"{category}{wash}{qty_tier}"


def _read_raw_apparel_styles(path: str) -> dict:
    """STYLE -> {"etd": date|None, "label": str|None}. Q=Style Code, AU=RDD(KR)/AX=RDD(CN)/BA=ETA(GL)
    중 최소값이 etd. D=item코드+AN=Qty TTL+BU=워시여부로 라벨(구분+워시기준+수량기준)도 직접 조립한다.
    AI(모델수)==1인 행만 그 스타일의 대표 행으로 세고, AS(상태)==DROP인 스타일은 뺀다.
    이 시트가 그 시즌의 전체 스타일 마스터라 FIT TRACK CHART/DUE_DATA(2)에 아직 없는 신규 스타일도 잡힌다."""
    with tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False) as tmp:
        shutil.copy2(path, tmp.name)
        wb = openpyxl.load_workbook(tmp.name, data_only=True, keep_vba=True)
    ws = wb[_RAW_SHEET]
    header_row = _find_header_row(ws, _RAW_COL_STYLE, "Style Code")
    _check_headers(ws, header_row, {
        _RAW_COL_STYLE: "Style Code",
        _RAW_COL_ITEM: "item",
        _RAW_COL_WASHED: "Washed",
        _RAW_COL_RDD_KR: "RDD (KR)",
        _RAW_COL_RDD_CN: "RDD (CN)",
        _RAW_COL_ETA_GL: "ETA (GL)",
    })
    _check_headers(ws, header_row - 1, {
        _RAW_COL_MODEL_COUNT: "모델수",
        _RAW_COL_QTY_TTL: "총 수량",
    })
    _check_headers(ws, 1, {
        _RAW_COL_STATUS: "status_godrop_TTL",
    })
    result = {}
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        style_code = r[_RAW_COL_STYLE - 1]
        if not style_code or not isinstance(style_code, str) or style_code in result:
            continue
        if r[_RAW_COL_MODEL_COUNT - 1] != 1:
            continue
        if str(r[_RAW_COL_STATUS - 1]).strip().upper() == "DROP":
            continue
        dates = [d for d in (
            _as_date(r[_RAW_COL_RDD_KR - 1]),
            _as_date(r[_RAW_COL_RDD_CN - 1]),
            _as_date(r[_RAW_COL_ETA_GL - 1]),
        ) if d]
        result[style_code] = {
            "etd": min(dates) if dates else None,
            "label": compute_label(r[_RAW_COL_ITEM - 1], r[_RAW_COL_WASHED - 1], r[_RAW_COL_QTY_TTL - 1]),
        }
    return result


def read_fit_track_raw(path: str, raw_apparel_path: str = None) -> list[dict]:
    """DUE_DATA(2) 값(offset 무관, 고정)과 offset 계산에 필요한 label/etd를 분리해서 그대로 반환.
    브라우저가 DUE DATE 설정 기준표를 바꿀 때마다 이 raw 데이터로 다시 계산할 수 있게 하기 위함.
    스타일 목록(모델수) 기준은 항상 FIT TRACK CHART(26FW FITTING 현황). raw_apparel_path 넘기면
    RAW(Apparel)에서 구분 라벨/ETD만 보충해서 DUE_DATA(2)에 없는 스타일의 due date 계산 정확도를 높인다."""
    # 사용자가 엑셀을 열어둔 채로 스케줄이 돌 수 있어 파일 잠금을 피하려고 임시 복사본을 읽는다.
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        shutil.copy2(path, tmp.name)
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
    due_data = _read_due_data_sheet(wb)
    fit_track = _read_fit_track_sheet(wb)
    raw_apparel = _read_raw_apparel_styles(raw_apparel_path) if raw_apparel_path else {}

    rows = []
    for style_code, track in fit_track.items():
        raw = raw_apparel.get(style_code, {"etd": None, "label": None})
        entry = due_data.get(style_code)
        label = raw["label"] or (entry["label"] if entry else None)
        due = entry["due"] if entry else {"FIT": None, "PP": None, "TOP": None}
        etd = raw["etd"] or track["etd"]

        rows.append({
            "style_code": style_code,
            "vendor": track["vendor"],
            "item": track["item"],
            "quarter": track["quarter"],
            "td": track["td"],
            "qa": track["qa"],
            "label": label,
            "etd": etd.isoformat() if etd else None,
            "fit_due": due["FIT"].isoformat() if due["FIT"] else None,
            "pp_due": due["PP"].isoformat() if due["PP"] else None,
            "top_due": due["TOP"].isoformat() if due["TOP"] else None,
            "fit_done": track["fit_done"],
            "pp_done": track["pp_done"],
            "top_done": track["top_done"],
            "detail": {
                stage: {
                    "round": d["round"],
                    "status": d["status"],
                    "confirm_date": d["confirm_date"].isoformat() if d["confirm_date"] else None,
                    "reason": d["reason"],
                    "first_received": d["first_received"].isoformat() if d["first_received"] else None,
                    "rounds": [
                        {
                            "round": r["round"],
                            "received": r["received"].isoformat() if r["received"] else None,
                            "status": r["status"],
                            "confirm_date": r["confirm_date"].isoformat() if r["confirm_date"] else None,
                        }
                        for r in d["rounds"]
                    ],
                }
                for stage, d in track["detail"].items()
            },
        })
    return rows


def _offset_lookup_label(label: str | None, label_offsets: dict) -> str | None:
    """기준표에 label이 그대로 없을 때의 폴백. DENIM은 기준표에 "워싱일반" 한 줄뿐이라(사용자 제공 표
    기준), DENIM인데 논워싱으로 찍힌 극소수 예외 건은 DENIM워싱일반 기준을 그대로 쓴다."""
    if label in label_offsets:
        return label
    if label and label.startswith("DENIM논워싱"):
        alt = "DENIM워싱" + label[len("DENIM논워싱"):]
        if alt in label_offsets:
            return alt
    return None


def resolve_due_dates(raw_rows: list[dict], label_offsets: dict = None) -> list[dict]:
    """read_fit_track_raw() 결과에 DUE DATE 설정 기준표(offset)를 적용해서 fit_due/pp_due/top_due를
    실제 date로 채운 행을 돌려준다. label_offsets 안 넘기면 LABEL_OFFSETS(기본표) 씀."""
    label_offsets = label_offsets or LABEL_OFFSETS
    rows = []
    for r in raw_rows:
        due = {
            "FIT": _as_date(datetime.fromisoformat(r["fit_due"])) if r["fit_due"] else None,
            "PP": _as_date(datetime.fromisoformat(r["pp_due"])) if r["pp_due"] else None,
            "TOP": _as_date(datetime.fromisoformat(r["top_due"])) if r["top_due"] else None,
        }
        missing = [stage for stage, d in due.items() if d is None]
        lookup_label = _offset_lookup_label(r["label"], label_offsets)
        if missing and lookup_label and r["etd"]:
            etd = date.fromisoformat(r["etd"])
            offsets = label_offsets[lookup_label]
            for stage in missing:
                due[stage] = etd - timedelta(days=offsets[stage])
        rows.append({
            "style_code": r["style_code"],
            "fit_due": due["FIT"],
            "pp_due": due["PP"],
            "top_due": due["TOP"],
            "fit_done": r["fit_done"],
            "pp_done": r["pp_done"],
            "top_done": r["top_done"],
        })
    return rows


def read_fit_track_rows(path: str, label_offsets: dict = None, raw_apparel_path: str = None) -> list[dict]:
    return resolve_due_dates(read_fit_track_raw(path, raw_apparel_path), label_offsets)


OWNER_BY_STAGE = {"FIT": "TD", "PP": "QA", "TOP": "QA"}


def compute_progress_from_xlsx(rows: list[dict], as_of_date: date) -> dict:
    result = {"TD": {}, "QA": {}}
    done_fields = {"FIT": "fit_done", "PP": "pp_done", "TOP": "top_done"}
    due_fields = {"FIT": "fit_due", "PP": "pp_due", "TOP": "top_due"}

    for row in rows:
        for stage in ("FIT", "PP", "TOP"):
            is_done = row[done_fields[stage]]
            due = row[due_fields[stage]]
            is_due = bool(due and due <= as_of_date)

            owner_type = OWNER_BY_STAGE[stage]
            bucket = result[owner_type].setdefault(
                stage, {"total_done": 0, "total_all": 0, "baseline_done": 0, "baseline_all": 0}
            )
            bucket["total_all"] += 1
            if is_done:
                bucket["total_done"] += 1
            if is_due:
                bucket["baseline_all"] += 1
                if is_done:
                    bucket["baseline_done"] += 1

    return result
